import time
import pandas as pd
import openpyxl
import math
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, fills
import glob
import os
import streamlit as st


from datetime import datetime
start_time = datetime.now()


def octant_analysis(xpath,ypath,mod):

    os.chdir(xpath)
    lst_files=glob.glob('*.xlsx')
    sat=0
    for file in lst_files :

        os.chdir(xpath) 
        DATA = pd.read_excel(file) 
        
        avg_u = DATA['U'].mean()  
        avg_v = DATA['V'].mean()
        avg_w = DATA['W'].mean()
        
        DATA["U_Avg"] = '' 
        DATA["V_Avg"] = ''
        DATA["W_Avg"] = ''
        
        DATA.iloc[0, 4] = round(avg_u,3)
        DATA.iloc[0, 5] = round(avg_v,3)
        DATA.iloc[0, 6] = round(avg_w,3)

        
        DATA["U'=U - U avg"] = round(DATA["U"]-avg_u,3)
        DATA["V'=V - V avg"] = round(DATA["V"]-avg_v,3)
        DATA["W'=W - W avg"] = round(DATA["W"]-avg_w,3)

       
        #**************Data PreProcessing*************#

        DATA["Octant"] = ''  

        l = len(DATA)  

        
        for i in range(0, l):

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+1"  # for +1

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-1"  # for -1

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+2"  # for +2

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-2"  # for -2

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+3"  # for +3

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-3"  # for -3

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+4"  # for +4

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-4"  # for -4

        #**************Octant Identification**************#

        # creating empty Column without header and assigned "User input" to row 3
        DATA[""] = "  "
        DATA[" "] = " "
        DATA.iloc[0, 12] = "Mod "+str(mod)

        
        DATA["Octant ID"] = " "
        DATA.loc[0, "Octant ID"] = "Overall Octant"

        
        oct_count = DATA['Octant'].value_counts()

        arr = ["+1", "-1", "+2", "-2", "+3", "-3",
            "+4", "-4"]  

        oct_cnt = {}  
        for i in range(8):
            s = arr[i]
            # appending the overall count of octant and octant value in dict 
            oct_cnt.update({s:oct_count[s]})
            # And assigning a count values to respectively Coloumns
            DATA.loc[0, s] = oct_count[s]

       

        #sorting the dict by keys
        sortedbyval = {k: v for k, v in sorted(oct_cnt.items(), key=lambda item: item[1])}
        #storing the sorted values in a list
        sortedbykey_lst = list(sortedbyval.keys())

        #Creating a dictionary for easy purpose
        octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                                "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}
        # creating empty columns
        DATA["Rank 1"] = ''  
        DATA["Rank 2"] = ''
        DATA["Rank 3"] = ''
        DATA["Rank 4"] = ''
        DATA["Rank 5"] = ''
        DATA["Rank 6"] = ''
        DATA["Rank 7"] = ''
        DATA["Rank 8"] = ''
        DATA["Rank1 Octant ID"] = " "
        #Creating another dictionary for referance
        dic_rank = {"+1": "Rank 1", "-1": "Rank 2", "+2": "Rank 3", "-2": "Rank 4",
                    "+3": "Rank 5", "-3": "Rank 6", "+4": "Rank 7", "-4": "Rank 8"}  
        
        #i=0
        for i in range(8):
            # appending the octant ranks of octants
            DATA.loc[0, dic_rank[sortedbykey_lst[i]]] = 8 - i  
            if (8-i == 1):
                # appending the highest rank octant and its corresponding octant name
                DATA.loc[0, "Rank1 Octant ID"] = sortedbykey_lst[i]
                DATA.loc[0, "Rank1 Octant Name"] = octant_name_id_mapping[str(int(DATA.loc[0, "Rank1 Octant ID"]))]

#***********Added Some Columns And Rows for MOD Count***************#

        x = 0  
        t = 1  

        count_rank_mod = [0]*8  
        while (x < l):
            # creating a dictionary for reference
            d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
                "-3": 5, "+4": 6, "-4": 7}  

            # count values of each octant is stored for MOD ranges
            oct_cnt_mod = [0]*8
             # for storing octant count as key and and coreesponding octant value as value in dict
            oct_cnt = {}  
            for i in range(x, x+mod, 1):

                if (i >= l):
                    break  
                s3 = DATA.at[i, "Octant"]
                # incrementing by one of count values of corresponding octants
                oct_cnt_mod[d1[s3]] += 1

            i=0
            for i in range(8):
                s = arr[i]
                # assigning overall count of octants in each interval
                DATA.loc[t, s] = oct_cnt_mod[i]
                # appending the overall count of octant and octant value in dict
                oct_cnt.update({s:oct_cnt_mod[i]})
                

           
            sortedbyval = {k: v for k, v in sorted(oct_cnt.items(), key=lambda item: item[1])}
            #storing the sorted values in a list
            sortedbykey_lst = list(sortedbyval.keys())

            i=0
            for i in range(8):
                # appending the octant ranks of octants
                DATA.loc[t, dic_rank[sortedbykey_lst[i]]] = 8 - i  
                if (8-i == 1):
                    # appending the highest rank octant and its corresponding octant name
                    DATA.loc[t, "Rank1 Octant ID"] = sortedbykey_lst[i]
                    DATA.loc[t, "Rank1 Octant Name"] = octant_name_id_mapping[str(int(DATA.loc[t, "Rank1 Octant ID"]))]
                    # incrementing by one of corresponding octant
                    count_rank_mod[d1[sortedbykey_lst[i]]] += 1

            if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
                DATA.loc[t, "Octant ID"] = str(x)+"-"+str(l-1) 
            else:
                DATA.loc[t, "Octant ID"] = str(x)+"-"+str(x+mod-1)

            x += mod
            t += 1

    #*********************Octant Count Based on Mod Values***********************#

        t += 3
        DATA.loc[t, "+1"] = "Octant ID"
        DATA.loc[t, "-1"] = "Octant Name"
        DATA.loc[t, "+2"] = "Count of Rank1 of Mod Values"
        t += 1
        i = 0
        # appending the octant ranks of octants
        for ID, name in octant_name_id_mapping.items():  
            # appending the Octant IDs, Octant Name ,and count of Rank1 of mod values
            DATA.loc[t, "+1"] = int(ID)
            DATA.loc[t, "-1"] = name
            DATA.loc[t, "+2"] = count_rank_mod[i]
            t += 1
            i += 1

    #*****************Octant Count Based on Mod Values*********************#
        DATA["  "] = ""
        DATA["   "] = " "
        DATA.iloc[0, 33] = "From"
        DATA.loc["Octant #"] = " "
        arr = [" +1", " -1", " +2", " -2", " +3", " -3", " +4", " -4"]

        j = 0
        for i in range(0, 8):
            # updating Octant ID column
            DATA.loc[i, "Octant #"] = arr[j]  
            j += 1

        j = 0
        for j in range(0, 8):
            # verifing the count of octants
            s1 = arr[j]  
            DATA[s1] = " "

        t1 = 0
        t2 = 1
        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
            "+3": 4, "-3": 5, "+4": 6, "-4": 7}
        d2 = {"+1": " +1", "-1": " -1", "+2": " +2", "-2": " -2",
            "+3": " +3", "-3": " -3", "+4": " +4", "-4": " -4"}

        while (1):
            if (t2 == l):
                break
            s1 = DATA.at[t1, "Octant"]  
            s2 = DATA.at[t2, "Octant"]  
            # checking if cell is empty/null
            if (DATA.loc[d1[s1], d2[s2]] == " "):  
                # adding one
                DATA.loc[d1[s1], d2[s2]] = 1  
            else:
                # increamenting the count by one and updating it to column
                DATA.loc[d1[s1], d2[s2]] = int(DATA.loc[d1[s1], d2[s2]]) + 1
            t1 += 1
            t2 += 1

        t = 7
        x = 0
        while (x < l):
            t += 4
            DATA.loc[t, "Octant #"] = "Mod Transition Count"
            # Writing MOD ranges in Octant ID Coloumn
            if ((x+mod) > l):  
                # for last index
                DATA.loc[t+1, "Octant #"] = str(x)+"-"+str(l-1)
            else:
                DATA.loc[t+1, "Octant #"] = str(x)+"-"+str(x+mod-1)
            DATA.loc[t+1, " +1"] = "To"
            t += 2
            arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
            DATA.loc[t, "Octant #"] = "Octant #"
            DATA.iloc[t+1, 33] = "From"
            # h stores column labels
            h = DATA.columns  
            # header name in index format(integer) (here ,y=13)
            y = h.get_loc(" +1")
            j = 0
            #Updating row and column
            for i in range(y, y+8): 
                DATA.iloc[t, i] = arr[j]
                j += 1

            j = 0
            for i in range(t+1, t+9):  
                DATA.loc[i, "Octant #"] = arr[j]
                j += 1
             # each interval
            for i in range(x, x+mod): 

                if (i == l-1):
                    break
                s1 = DATA.at[i, "Octant"]  
                s2 = DATA.at[i+1, "Octant"]  

                if (DATA.loc[t+d1[s1]+1, d2[s2]] == " "):  
                    # checking if cell is empty/null
                    DATA.loc[t+d1[s1]+1, d2[s2]] = 1  
                    # adding one
                else:
                    # increamenting the count by one and updating it to coloumn
                    DATA.loc[t+d1[s1]+1, d2[s2]
                            ] = int(DATA.loc[t+d1[s1]+1, d2[s2]]) + 1
            t += 8
            x += mod

        #**************************#
        DATA["    "] = " "
        # Creating empty column with Octant as a header
        DATA["Octant ##"] = " "  
        arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        for i in range(8):
            # appending values in octant column
            DATA.loc[i, "Octant ##"] = arr[i]

        DATA["Longest Subsequence Length"] = " "
        DATA["Count"] = " "

        x = 0
        # Longest subsequence length for respectively octant value
        #initlizing a max_count with all zeroes
        max_count = [0]*8

        # for count of LSL for respectively octant values
        #initlizing a LSL_count with all zeroes
        LSL_count = [0]*8
        # creating a dictionary for referance
        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
            "-3": 5, "+4": 6, "-4": 7}  
        # Creating an empty 2d list of size of 8
        # where each list stores the upper range(Time Range) value of thier respectively Octants
        time_range = []
        for i in range(8):
            time_range.append([])

        while (x < l):
            s1 = DATA.at[x, "Octant"]
            count = 0
            j = x
            # counting length of sequence
            while (1):  
                # breaking if next element is not equal to s1
                if (j >= l or DATA.at[j, "Octant"] != s1):
                    break
                count += 1
                j += 1

            x += count
            temp = max_count[d1[s1]]

            # updating a maximum count of value if current count is greater the current max
            max_count[d1[s1]] = max(max_count[d1[s1]], count)

            if (count > temp):
                # Reassigning the values of LSL count to one
                LSL_count[d1[s1]] = 1
                # if list is empty appending  Upper range Value
                if (len(time_range[d1[s1]]) == 0):
                    time_range[d1[s1]].append(j-1)

                else:
                    time_range[d1[s1]].clear()  
                    # appending a curent upper range value to the same clered octant list
                    time_range[d1[s1]].append(j-1)

            if (count == temp):
                # incremneting the count of LSL by one
                LSL_count[d1[s1]] += 1
                # appending to the pre-existing(non-empty) list having same LSL of respective Octant
                time_range[d1[s1]].append(j-1)

        for i in range(8):
            # updating Longest subsequence length for respectively octant values
            DATA.loc[i, "Longest Subsequence Length"] = max_count[i]

        max_l_cnt=0
        for j in range(8):
            # updating count of LSL for respectively octant values
            DATA.loc[j, "Count"] = LSL_count[j]
            max_l_cnt+=LSL_count[j]

        DATA["     "] = " "  # Empty Column without Header
        DATA["Octant ####"] = " "  # Empty Column
        DATA[" Longest Subsequence Length"] = " "  # Empty Column
        DATA[" Count"] = " "  # Empty Column
       

        t = 0  # row pointer
        for i in range(8):
            DATA.loc[t, "Octant ####"] = arr[i]  # Updating Octant Values
            # Updating LSL of Octants
            DATA.loc[t, " Longest Subsequence Length"] = max_count[i]
            # updating count of LSl of Octants
            DATA.loc[t, " Count"] = LSL_count[i]
            t += 1  # t points to next row
            DATA.loc[t, "Octant ####"] = "Time"
            DATA.loc[t, " Longest Subsequence Length"] = "From"
            DATA.loc[t, " Count"] = "To"

            t += 1  # t points to next row
            for j in range(LSL_count[i]):
                # Appending lower range # From
                DATA.loc[t, " Longest Subsequence Length"] = 0.01 * \
                    ((time_range[d1[arr[i]]][j])-(max_count[i]-1))
                # Appending Upper range #To
                DATA.loc[t, " Count"] = 0.01*time_range[d1[arr[i]]][j]
                t += 1

        inp= lst_files[sat].replace('.xlsx'," cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")        
        
        os.chdir(ypath)
        print(inp)        
        DATA.to_excel(inp, index=False)
        wb=openpyxl.load_workbook(inp)
        ws=wb['Sheet1']
        thin_border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
                

        fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')

        #define size of the table 
        row_num=math.ceil(l/mod)+2
        col_num=19 
        #location of the Table 
        row_loc=1
        col_loc=14


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border
                if((ws.cell(row=i,column=j).value == 1)):
                    ws.cell(row=i, column=j).fill=fill_cell
              

        #define size of the table 
        row_num=9
        col_num=3
        #location of the Table 
        row_loc=math.ceil(l/mod)+6
        col_loc=15


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border
       
        #define size of the table 
        row_num=9
        col_num=3
        #location of the Table 
        row_loc=1
        col_loc=45


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border

        #define size of the table 
        row_num=9
        col_num=9 
        #location of the Table 
        row_loc=1
        col_loc=35

        #Number of Tables 
        Table_num=math.ceil(l/mod)+1
        dis=5 # distance between the tables 

        for _ in range(Table_num):
            k=0
            for i in range (row_loc,row_loc+row_num):
                    
                    if(i>row_loc):
                         ws.cell(row=i, column=col_loc+k).fill=fill_cell
                    for j in range (col_loc,col_num+col_loc):
                        ws.cell(row=i, column=j).border=thin_border
                       
                    k+=1
                        
            row_loc= row_loc+row_num+dis

        row_num=l
        col_num=3
        #location of the Table 
        row_loc=1
        col_loc=49


        for i in range (row_loc,row_loc+row_num): # for mod
            # print(ws.cell(row=i, column=50).value)
            if(ws.cell(row=i, column=50).value == " "):
                break
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border            


        sat=sat+1
        wb.save(inp)

def octant_analysis_1(NAME,DATA,mod):

     # reading the input file
     # Calculating average of U,V,W using mean functions
        avg_u = DATA['U'].mean()  
        avg_v = DATA['V'].mean()
        avg_w = DATA['W'].mean()
        # Creating average for coloumns U,V,W
        DATA["U_Avg"] = ''  
        DATA["V_Avg"] = ''
        DATA["W_Avg"] = ''
        # assigning the values to respectivley Coloumn
        DATA.iloc[0, 4] = round(avg_u,3)
        DATA.iloc[0, 5] = round(avg_v,3)
        DATA.iloc[0, 6] = round(avg_w,3)

        # Creating new coloumns with Header U',V',W'
        DATA["U'=U - U avg"] = round(DATA["U"]-avg_u,3)
        DATA["V'=V - V avg"] = round(DATA["V"]-avg_v,3)
        DATA["W'=W - W avg"] = round(DATA["W"]-avg_w,3)

        # DATA.to_xlsx('octant_output.xlsx')

        #######          Data PreProcessing     ###########

        DATA["Octant"] = ''  # Creatig a empty Column with Header as Octant

        l = len(DATA)  # length of DataFrame = 29745

        # creating octant column ,and Identifying the octant value for each triple(U_1,V_1,W_1)
        for i in range(0, l):

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+1"  # for +1

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-1"  # for -1

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+2"  # for +2

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] >= 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-2"  # for -2

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+3"  # for +3

            if (DATA.loc[i, "U'=U - U avg"] < 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-3"  # for -3

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] >= 0):
                DATA.loc[i, "Octant"] = "+4"  # for +4

            if (DATA.loc[i, "U'=U - U avg"] >= 0 and DATA.loc[i, "V'=V - V avg"] < 0 and DATA.loc[i, "W'=W - W avg"] < 0):
                DATA.loc[i, "Octant"] = "-4"  # for -4

                ######  Octant Identification  ########

        # creating empty Column without header and assigned "User input" to row 3
        DATA[""] = "  "
        DATA[" "] = " "
        DATA.iloc[0, 12] = "Mod "+str(mod)

        # creating a Coloumn with header as Octant ID
        DATA["Octant ID"] = " "
        DATA.loc[0, "Octant ID"] = "Overall Octant"

        # oct_count stores a count of unique elements i.e. count of +1,-1,+2,-2,+3,-4,+4
        oct_count = DATA['Octant'].value_counts()

        arr = ["+1", "-1", "+2", "-2", "+3", "-3",
            "+4", "-4"]  # cretaed for reference

        oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
        for i in range(8):
            s = arr[i]
            # appending the overall count of octant and octant value in dict i.e for Ex:(2610,"+1")
            oct_cnt.update({s:oct_count[s]})
            # And assigning a count values to respectively Coloumns
            DATA.loc[0, s] = oct_count[s]

       
        #sorting the dict by keys
        sortedbyval = {k: v for k, v in sorted(oct_cnt.items(), key=lambda item: item[1])}
        #storing the sorted values in a list
        sortedbykey_lst = list(sortedbyval.keys())

        octant_name_id_mapping = {"1": "Internal outward interaction", "-1": "External outward interaction", "2": "External Ejection",
                                "-2": "Internal Ejection", "3": "External inward interaction", "-3": "Internal inward interaction", "4": "Internal sweep", "-4": "External sweep"}

        DATA["Rank 1"] = ''  # created empty columns
        DATA["Rank 2"] = ''
        DATA["Rank 3"] = ''
        DATA["Rank 4"] = ''
        DATA["Rank 5"] = ''
        DATA["Rank 6"] = ''
        DATA["Rank 7"] = ''
        DATA["Rank 8"] = ''
        DATA["Rank1 Octant ID"] = " "

        dic_rank = {"+1": "Rank 1", "-1": "Rank 2", "+2": "Rank 3", "-2": "Rank 4",
                    "+3": "Rank 5", "-3": "Rank 6", "+4": "Rank 7", "-4": "Rank 8"}  # for reference
        
        #i=0
        for i in range(8):
            DATA.loc[0, dic_rank[sortedbykey_lst[i]]] = 8 - i  # appending the octant ranks of octants
            if (8-i == 1):
                # appending the highest rank octant and its corresponding octant name
                DATA.loc[0, "Rank1 Octant ID"] = sortedbykey_lst[i]
                DATA.loc[0, "Rank1 Octant Name"] = octant_name_id_mapping[str(int(DATA.loc[0, "Rank1 Octant ID"]))]

                ###########   Added Some Columns And Rows for MOD Count   ##########

        x = 0  # for findind octant values for MOD ranges
        t = 1  # for row pointer

        count_rank_mod = [0]*8  # Count of rank mod values
        while (x < l):

            d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
                "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary for reference

            # count values of each octant is stored for MOD ranges
            oct_cnt_mod = [0]*8

            oct_cnt = {}  # for storing octant count as key and and coreesponding octant value as value in dict
            for i in range(x, x+mod, 1):

                if (i >= l):
                    break  # bound check
                s3 = DATA.at[i, "Octant"]
                # incrementing by one of count values of corresponding octants
                oct_cnt_mod[d1[s3]] += 1

            i=0
            for i in range(8):
                s = arr[i]
                # assigning overall count of octants in each interval
                DATA.loc[t, s] = oct_cnt_mod[i]
                # appending the overall count of octant and octant value in dict
                oct_cnt.update({s:oct_cnt_mod[i]})
               
        
            #sorting the dict by keys
            sortedbyval = {k: v for k, v in sorted(oct_cnt.items(), key=lambda item: item[1])}
            #storing the sorted values in a list
            sortedbykey_lst = list(sortedbyval.keys())

            i=0
            for i in range(8):
                DATA.loc[t, dic_rank[sortedbykey_lst[i]]] = 8 - i  # appending the octant ranks of octants
                if (8-i == 1):
                    # appending the highest rank octant and its corresponding octant name
                    DATA.loc[t, "Rank1 Octant ID"] = sortedbykey_lst[i]
                    DATA.loc[t, "Rank1 Octant Name"] = octant_name_id_mapping[str(int(DATA.loc[t, "Rank1 Octant ID"]))]
                    # incrementing by one of corresponding octant
                    count_rank_mod[d1[sortedbykey_lst[i]]] += 1

            if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
                DATA.loc[t, "Octant ID"] = str(x)+"-"+str(l-1)  # for last index(i.e) 2744
            else:
                DATA.loc[t, "Octant ID"] = str(x)+"-"+str(x+mod-1)

            x += mod
            t += 1

    #*****************Octant Count Based on Mod Values*********************#

        t += 3
        DATA.loc[t, "+1"] = "Octant ID"
        DATA.loc[t, "-1"] = "Octant Name"
        DATA.loc[t, "+2"] = "Count of Rank1 of Mod Values"
        t += 1
        i = 0
        for ID, name in octant_name_id_mapping.items():  # iterating through a dict
            # appending the Octant IDs, Octant Name ,and count of Rank1 of mod values
            DATA.loc[t, "+1"] = int(ID)
            DATA.loc[t, "-1"] = name
            DATA.loc[t, "+2"] = count_rank_mod[i]
            t += 1
            i += 1

    #*******************Octant Count Based on Mod Value*********************#
        DATA["  "] = ""
        DATA["   "] = " "
        DATA.iloc[0, 33] = "From"
        DATA.loc["Octant #"] = " "
        arr = [" +1", " -1", " +2", " -2", " +3", " -3", " +4", " -4"]
        # h = DATA.columns  # h stores column labels
        # y = h.get_loc("+1")  # header name in index format(integer) (here ,y=13)

        j = 0
        for i in range(0, 8):
            # updating Octant ID column
            DATA.loc[i, "Octant #"] = arr[j]  
            j += 1

        j = 0
        for j in range(0, 8):
            # verifing the count of octants
            s1 = arr[j]  
            DATA[s1] = " "

        t1 = 0
        t2 = 1
        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3,
            "+3": 4, "-3": 5, "+4": 6, "-4": 7}
        d2 = {"+1": " +1", "-1": " -1", "+2": " +2", "-2": " -2",
            "+3": " +3", "-3": " -3", "+4": " +4", "-4": " -4"}

        while (1):
            if (t2 == l):
                break
            s1 = DATA.at[t1, "Octant"]  # From
            s2 = DATA.at[t2, "Octant"]  # To
            # print(DATA.loc[d1[s1], d2[s2]])
            if (DATA.loc[d1[s1], d2[s2]] == " "):  # checking if cell is empty/null
                DATA.loc[d1[s1], d2[s2]] = 1  # adding one
            else:
                # increamenting the count by one and updating it to coloumn
                DATA.loc[d1[s1], d2[s2]] = int(DATA.loc[d1[s1], d2[s2]]) + 1
            t1 += 1
            t2 += 1

        t = 7
        x = 0
        while (x < l):
            t += 4
            DATA.loc[t, "Octant #"] = "Mod Transition Count"
            if ((x+mod) > l):  # Writing MOD ranges in Octant ID Coloumn
                # for last index(i.e) 2744
                DATA.loc[t+1, "Octant #"] = str(x)+"-"+str(l-1)
            else:
                DATA.loc[t+1, "Octant #"] = str(x)+"-"+str(x+mod-1)
            DATA.loc[t+1, " +1"] = "To"
            t += 2
            arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
            DATA.loc[t, "Octant #"] = "Octant #"
            DATA.iloc[t+1, 33] = "From"
            h = DATA.columns  # h stores column labels
            # header name in index format(integer) (here ,y=13)
            y = h.get_loc(" +1")
            j = 0
            for i in range(y, y+8):  # updating a row
                DATA.iloc[t, i] = arr[j]
                j += 1

            j = 0
            for i in range(t+1, t+9):  # updating Coloumn
                DATA.loc[i, "Octant #"] = arr[j]
                j += 1

            for i in range(x, x+mod):  # each interval

                if (i == l-1):
                    break
                s1 = DATA.at[i, "Octant"]  # From
                s2 = DATA.at[i+1, "Octant"]  # To

                if (DATA.loc[t+d1[s1]+1, d2[s2]] == " "):  # checking if cell is empty/null
                    DATA.loc[t+d1[s1]+1, d2[s2]] = 1  # adding one
                else:
                    # increamenting the count by one and updating it to coloumn
                    DATA.loc[t+d1[s1]+1, d2[s2]
                            ] = int(DATA.loc[t+d1[s1]+1, d2[s2]]) + 1
            t += 8
            x += mod

                                    ##################
        DATA["    "] = " "
        DATA["Octant ##"] = " "  # Creating empty column with Octant as a header
        arr = ["+1", "-1", "+2", "-2", "+3", "-3", "+4", "-4"]
        for i in range(8):
            # appending values in octant column
            DATA.loc[i, "Octant ##"] = arr[i]

        DATA["Longest Subsequence Length"] = " "
        DATA["Count"] = " "

        x = 0
        # Longest subsequence length for respectively octant values #initlizing a max_count with all zeroes  #initlizing a max_count with all zeroes
        max_count = [0]*8

        # for count of LSL for respectively octant values # initlizing a max_count with all zeroes #initlizing a LSL_count with all zeroes
        LSL_count = [0]*8
        d1 = {"+1": 0, "-1": 1, "+2": 2, "-2": 3, "+3": 4,
            "-3": 5, "+4": 6, "-4": 7}  # creating a dictionary

        # Creating an empty 2d list of size of 8
        # where each list stores the upper range(Time Range) value of thier respectively Octants
        time_range = []
        for i in range(8):
            time_range.append([])

        while (x < l):
            s1 = DATA.at[x, "Octant"]
            count = 0
            j = x
            while (1):  # counting length of sequence
                # breaking if next element is not equal to s1
                if (j >= l or DATA.at[j, "Octant"] != s1):
                    break
                count += 1
                j += 1

            x += count
            temp = max_count[d1[s1]]

            # updating a maximum count of value if current count is greater the current max
            max_count[d1[s1]] = max(max_count[d1[s1]], count)

            if (count > temp):
                # Reassigning the values of LSL count to one
                LSL_count[d1[s1]] = 1
                # if list is empty appending  Upper range Value
                if (len(time_range[d1[s1]]) == 0):
                    time_range[d1[s1]].append(j-1)

                else:
                    time_range[d1[s1]].clear()  # Clearing the list
                    # appending a curent upper range value to the same clered octant list
                    time_range[d1[s1]].append(j-1)

            if (count == temp):
                # incremneting the count of LSL by one
                LSL_count[d1[s1]] += 1
                # appending to the pre-existing(non-empty) list having same LSL of respective Octant
                time_range[d1[s1]].append(j-1)

        for i in range(8):
            # updating Longest subsequence length for respectively octant values
            DATA.loc[i, "Longest Subsequence Length"] = max_count[i]

        max_l_cnt=0
        for j in range(8):
            # updating count of LSL for respectively octant values
            DATA.loc[j, "Count"] = LSL_count[j]
            max_l_cnt+=LSL_count[j]

        DATA["     "] = " "  # Empty Column without Header
        DATA["Octant ####"] = " "  # Empty Column
        DATA[" Longest Subsequence Length"] = " "  # Empty Column
        DATA[" Count"] = " "  # Empty Column
        
        t = 0  # row pointer
        for i in range(8):
            DATA.loc[t, "Octant ####"] = arr[i]  # Updating Octant Values
            # Updating LSL of Octants
            DATA.loc[t, " Longest Subsequence Length"] = max_count[i]
            # updating count of LSl of Octants
            DATA.loc[t, " Count"] = LSL_count[i]
            t += 1  # t points to next row
            DATA.loc[t, "Octant ####"] = "Time"
            DATA.loc[t, " Longest Subsequence Length"] = "From"
            DATA.loc[t, " Count"] = "To"

            t += 1  # t points to next row
            for j in range(LSL_count[i]):
                # Appending lower range # From
                DATA.loc[t, " Longest Subsequence Length"] = 0.01 * \
                    ((time_range[d1[arr[i]]][j])-(max_count[i]-1))
                # Appending Upper range #To
                DATA.loc[t, " Count"] = 0.01*time_range[d1[arr[i]]][j]
                t += 1

        inp= NAME.replace('.xlsx'," cm_vel_octant_analysis_mod_"+str(mod)+".xlsx")        
        
        os.chdir(r'C:\Users\Admin\Documents\GitHub\2001CB62_2022\proj2\output')
        print(inp)        
        DATA.to_excel(inp, index=False)
        wb=openpyxl.load_workbook(inp)
        ws=wb['Sheet1']
        thin_border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000')
                )
                

        fill_cell = PatternFill(fill_type=fills.FILL_SOLID,start_color='00FFFF00',end_color='00FFFF00')

        #define size of the table 
        row_num=math.ceil(l/mod)+2
        col_num=19 
        #location of the Table 
        row_loc=1
        col_loc=14


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border
                if((ws.cell(row=i,column=j).value == 1)):
                    ws.cell(row=i, column=j).fill=fill_cell
                
        row_num=9
        col_num=3
        #location of the Table 
        row_loc=math.ceil(l/mod)+6
        col_loc=15


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border
                
        #define size of the table 
        row_num=9
        col_num=3
        #location of the Table 
        row_loc=1
        col_loc=45


        for i in range (row_loc,row_loc+row_num): # for mod
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border

        #define size of the table 
        row_num=9
        col_num=9 
        #location of the Table 
        row_loc=1
        col_loc=35

        #Number of Tables 
        Table_num=math.ceil(l/mod)+1
        dis=5 # distance between the tables 

        for _ in range(Table_num):
            k=0
            for i in range (row_loc,row_loc+row_num):
                    
                    if(i>row_loc):
                         ws.cell(row=i, column=col_loc+k).fill=fill_cell
                    for j in range (col_loc,col_num+col_loc):
                        ws.cell(row=i, column=j).border=thin_border
                      
                    k+=1
                        
            row_loc= row_loc+row_num+dis

        row_num=l
        col_num=3
        #location of the Table 
        row_loc=1
        col_loc=49


        for i in range (row_loc,row_loc+row_num): # for mod
            
            if(ws.cell(row=i, column=50).value == " "):
                break
            for j in range (col_loc,col_num+col_loc):
                ws.cell(row=i, column=j).border=thin_border            


        wb.save(inp)   
st.set_page_config(page_title="Python CS_384 Project_2",page_icon=":four_leaf_clover:",layout="wide")     
new_title = '<p style="font-family:roboto; color:red; font-size: 50px;text-align:center ">Python Project 2</p>'
st.markdown(new_title, unsafe_allow_html=True)

z = st.sidebar.selectbox('SELECT INPUT TYPE',['NONE','FILE','FOLDER'])

if z=='FOLDER':
    xpath=st.text_input('ENTER INPUT FILE PATH:')
    mod=st.number_input('ENTER MOD VALUE:',value=50,step=1)
    ypath=st.text_input('ENTER DOWNLOAD LOCATION/PATH :')
    if st.button('RUN'):
        st.balloons()
        octant_analysis(xpath,ypath,mod) 
        st.success("Successfully completed!")


elif z=='FILE':
    
    sub_title = '<p style="font-family:roboto; color:Black; font-size: 30px">INPUT FILE</p>'
    st.markdown(sub_title, unsafe_allow_html=True)
    data_file = st.file_uploader("**UPLOAD INPUT FILE**",type=['xlsx'])
    if data_file is not None:
       file_details = {"Filename":data_file.name,"FileType":data_file.type,"FileSize":data_file.size}
       st.write(file_details)
       df = pd.read_excel(data_file)
    
       mod=st.number_input('ENTER MOD VALUE',value=22,step=1)
       if mod:
        if st.button('RUN'):
            octant_analysis_1(data_file.name,df,mod)
            bar = st.progress(2)             #progress bar
            for i in range(100):
                time.sleep(0.01)
                bar.progress(i + 1)
            with st.spinner('Running...'):
                time.sleep(2)
            st.success("Successfully completed!")
			              

end_time = datetime.now()

print('Duration of Program Execution: {}'.format(end_time - start_time))